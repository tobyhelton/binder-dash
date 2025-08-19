# -*- coding: utf-8 -*-
"""
Created on Tue Aug 12 17:32:43 2025

@author: tobyr
"""

import time
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

EDGE_DRIVER_PATH = r"C:\Users\tobyr\Downloads\edgedriver_win64\msedgedriver.exe"

def get_price_and_change_tcgplayer(driver, url):
    driver.get(url)
    time.sleep(2)

    # Price
    try:
        price_el = driver.find_element(By.CSS_SELECTOR, "span.price-points__upper__price")
        price = float(price_el.text.strip().replace('$', '').replace(',', ''))
    except:
        price = None

    # % Change
    try:
        change_el = driver.find_element(By.CSS_SELECTOR, "span.charts-positive.charts-change, span.charts-negative.charts-change")
        change_text = change_el.text.strip().replace('(', '').replace(')', '').replace('%', '').replace(',', '')
        percent_change = float(change_text)
    except:
        percent_change = None

    return price, percent_change


def get_price_and_change_collectr(driver, url):
    driver.get(url)
    time.sleep(2)

    # Price
    try:
        price_el = driver.find_element(By.CSS_SELECTOR, "h3.ml-2.font-bold.dark\\:text-textColorDark.text-secondaryText.text-md")
        price = float(price_el.text.strip().replace('$', '').replace(',', ''))
    except:
        price = None

    # Recent Change
    try:
        change_el = driver.find_element(By.CSS_SELECTOR, "span.mr-1")
        recent_change = float(change_el.text.strip().replace('+', ''))
    except:
        recent_change = None

    return price, recent_change


def get_price_pricecharting(driver, url):
    driver.get(url)
    time.sleep(2)

    try:
        price_el = driver.find_element(By.CSS_SELECTOR, "span.price.js-price")
        price = float(price_el.text.strip().replace('$', '').replace(',', ''))
    except:
        price = None

    return price


def update_excel(file_path):
    service = Service(EDGE_DRIVER_PATH)
    driver = webdriver.Edge(service=service)

    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        # --- TCGPlayer ---
        tcg_link_cell = row[6]  # Column G
        tcg_price_cell = row[7] # Column H
        tcg_change_cell = row[13] # Column N

        if tcg_link_cell.hyperlink:
            tcg_url = tcg_link_cell.hyperlink.target
            price, pct_change = get_price_and_change_tcgplayer(driver, tcg_url)
            tcg_price_cell.value = price if price is not None else "Not Found"
            tcg_change_cell.value = pct_change if pct_change is not None else "Not Found"
        else:
            tcg_price_cell.value = "No Link"
            tcg_change_cell.value = "No Link"

        # --- Collectr ---
        collectr_link_cell = row[8] # Column I
        collectr_price_cell = row[9] # Column J
        collectr_change_cell = row[14] # Column O

        if collectr_link_cell.hyperlink:
            collectr_url = collectr_link_cell.hyperlink.target
            price, recent_change = get_price_and_change_collectr(driver, collectr_url)
            collectr_price_cell.value = price if price is not None else "Not Found"
            collectr_change_cell.value = recent_change if recent_change is not None else "Not Found"
        else:
            collectr_price_cell.value = "No Link"
            collectr_change_cell.value = "No Link"

        # --- PriceCharting ---
        pc_link_cell = row[10] # Column K
        pc_price_cell = row[11] # Column L

        if pc_link_cell.hyperlink:
            pc_url = pc_link_cell.hyperlink.target
            price = get_price_pricecharting(driver, pc_url)
            pc_price_cell.value = price if price is not None else "Not Found"
        else:
            pc_price_cell.value = "No Link"

    wb.save(file_path)
    wb.close()
    driver.quit()
    print("Excel file updated successfully.")


# Run the update
update_excel(r"C:\Users\tobyr\OneDrive\Desktop\PersonalProj\ThePrizedCollecton2.xlsx")

